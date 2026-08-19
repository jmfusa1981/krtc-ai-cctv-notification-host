from __future__ import annotations

import csv
from datetime import datetime, time, timedelta
from io import BytesIO
from urllib.parse import urlencode

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, QuerySet
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from apps.cameras.models import Camera
from apps.events.models import Event, EventRecordingEvidence
from apps.ai_bridge.models import InferenceHost
from apps.events.services.snapshot_localizer import local_snapshot_url


EVENT_EXPORT_HEADERS = [
    "事件編號",
    "發生時間",
    "事件類型",
    "攝影機編號",
    "區域",
    "處理狀態",
    "AI 模型",
    "來源推論主機",
    "事件說明",
    "快照",
    "錄影證據狀態",
    "錄影證據時間窗",
    "錄影下載",
]


def _parse_local_datetime(value: str):
    if not value:
        return None

    parsed = parse_datetime(value)
    if parsed is None:
        try:
            parsed = datetime.strptime(value, "%Y-%m-%dT%H:%M")
        except ValueError:
            return None

    if timezone.is_naive(parsed):
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    return parsed


# KRTC V5.14 DAILY EVENT NUMBER FILTER
def _apply_record_number_filter(queryset: QuerySet[Event], value: str) -> QuerySet[Event]:
    """Filter MMDDXXXX display IDs while retaining legacy PK lookup as fallback."""
    digits = "".join(ch for ch in (value or "") if ch.isdigit())
    if len(digits) == 8:
        month = int(digits[:2])
        day = int(digits[2:4])
        sequence = int(digits[4:])
        try:
            datetime(2000, month, day)
        except ValueError:
            return queryset.none()
        return queryset.filter(
            record_date__month=month,
            record_date__day=day,
            record_sequence=sequence,
        )
    if digits:
        return queryset.filter(id=int(digits))
    return queryset


def _event_queryset(request: HttpRequest) -> QuerySet[Event]:
    # Show the current local calendar day plus the preceding 29 days.
    # This keeps the record window aligned with the PAO daily event-number date.
    local_today = timezone.localdate()
    start_date = local_today - timedelta(days=29)
    start_at = timezone.make_aware(
        datetime.combine(start_date, time.min),
        timezone.get_current_timezone(),
    )

    queryset = (
        Event.objects.select_related("camera", "ai_model")
        .filter(detected_at__gte=start_at)
        .order_by("-detected_at", "-id")
    )

    event_id = request.GET.get("event_id", "").strip()
    event_type = request.GET.get("event_type", "").strip()
    area = request.GET.get("area", "").strip()
    source_host = request.GET.get("source_host", "").strip()

    if event_id:
        queryset = _apply_record_number_filter(queryset, event_id)
    if event_type:
        queryset = queryset.filter(event_type=event_type)
    if area:
        queryset = queryset.filter(camera__area=area)
    if source_host:
        queryset = queryset.filter(
            Q(inference_host_code=source_host) | Q(source_host=source_host)
        )

    return queryset


def _inference_host_maps() -> tuple[dict[str, str], dict[str, str]]:
    """Return display-name maps keyed by host code and normalized base URL."""
    by_code: dict[str, str] = {}
    by_url: dict[str, str] = {}
    for host in InferenceHost.objects.all().only("host_code", "name", "base_url"):
        display_name = (host.name or host.host_code).strip()
        by_code[host.host_code] = display_name
        by_url[host.normalized_base_url] = display_name
    return by_code, by_url


def _inference_host_display(
    event: Event,
    by_code: dict[str, str],
    by_url: dict[str, str],
) -> str:
    if event.inference_host_code and event.inference_host_code in by_code:
        return by_code[event.inference_host_code]
    source_host = (event.source_host or "").rstrip("/")
    if source_host and source_host in by_url:
        return by_url[source_host]
    return event.inference_host_code or event.source_host or "—"


def _snapshot_url(request: HttpRequest, event: Event) -> str:
    local_url = local_snapshot_url(event)
    return request.build_absolute_uri(local_url) if local_url else ""


def _latest_recording(event: Event) -> EventRecordingEvidence | None:
    cached = getattr(event, "_prefetched_objects_cache", {})
    if "recording_evidences" in cached:
        evidences = cached["recording_evidences"]
        return evidences[0] if evidences else None
    return event.recording_evidences.order_by("-created_at").first()


def _recording_download_url(request: HttpRequest, evidence: EventRecordingEvidence | None) -> str:
    if evidence is None:
        return ""
    if evidence.file:
        try:
            return request.build_absolute_uri(evidence.file.url)
        except (ValueError, AttributeError):
            return evidence.file.url
    return evidence.download_url or ""


def _recording_window(evidence: EventRecordingEvidence | None) -> str:
    if evidence is None:
        return ""
    start_at = timezone.localtime(evidence.evidence_start_at).strftime("%Y-%m-%d %H:%M:%S")
    end_at = timezone.localtime(evidence.evidence_end_at).strftime("%Y-%m-%d %H:%M:%S")
    return f"{start_at} ~ {end_at}"


def _recording_local_mp4_ready(evidence: EventRecordingEvidence | None) -> bool:
    if evidence is None:
        return False
    if evidence.export_status != EventRecordingEvidence.STATUS_COMPLETED:
        return False
    if not evidence.file or not evidence.file.name:
        return False
    if not evidence.file.name.lower().endswith(".mp4"):
        return False
    try:
        return evidence.file.storage.exists(evidence.file.name)
    except Exception:
        return False


def _recording_status_display(evidence: EventRecordingEvidence | None) -> str:
    if evidence is None:
        return "\u5c1a\u672a\u8981\u6c42"
    if _recording_local_mp4_ready(evidence):
        return "\u53ef\u64ad\u653e"
    if evidence.export_status in {
        EventRecordingEvidence.STATUS_PENDING,
        EventRecordingEvidence.STATUS_REQUESTED,
    }:
        return "\u7b49\u5f85 NVR"
    if evidence.export_status == EventRecordingEvidence.STATUS_EXPORTING:
        return "\u8655\u7406\u4e2d"
    if evidence.export_status == EventRecordingEvidence.STATUS_COMPLETED:
        return "\u7b49\u5f85\u672c\u5730 MP4"
    if evidence.export_status == EventRecordingEvidence.STATUS_FAILED:
        return "\u53d6\u5f97\u5931\u6557"
    return evidence.get_export_status_display()


def _recording_ui_status(evidence: EventRecordingEvidence | None) -> str:
    if evidence is None:
        return "not-required"
    if _recording_local_mp4_ready(evidence):
        return "ready"
    if evidence.export_status in {
        EventRecordingEvidence.STATUS_PENDING,
        EventRecordingEvidence.STATUS_REQUESTED,
    }:
        return "waiting"
    if evidence.export_status in {
        EventRecordingEvidence.STATUS_EXPORTING,
        EventRecordingEvidence.STATUS_COMPLETED,
    }:
        return "processing"
    if evidence.export_status == EventRecordingEvidence.STATUS_FAILED:
        return "failed"
    return "processing"


def _recording_local_path(evidence: EventRecordingEvidence | None) -> str:
    if evidence is None or not evidence.file:
        return ""
    name = (evidence.file.name or "").replace("\\", "/")
    return f"media/{name}" if name else ""


def _event_row(request: HttpRequest, event: Event) -> list[str]:
    detected_at = timezone.localtime(event.detected_at).strftime("%Y-%m-%d %H:%M:%S")
    ai_model = str(event.ai_model) if event.ai_model else ""
    evidence = _latest_recording(event)
    return [
        event.record_number,
        detected_at,
        event.get_event_type_display(),
        event.camera_code or (event.camera.camera_code if event.camera else ""),
        event.camera.area if event.camera else "",
        event.get_status_display(),
        ai_model,
        event.source_host or "",
        event.description or "",
        _snapshot_url(request, event),
        evidence.get_export_status_display() if evidence else "未建立",
        _recording_window(evidence),
        _recording_download_url(request, evidence),
    ]


def _filter_context(request: HttpRequest) -> dict:
    return {
        "selected_event_id": request.GET.get("event_id", ""),
        "selected_event_type": request.GET.get("event_type", ""),
        "selected_area": request.GET.get("area", ""),
        "selected_source_host": request.GET.get("source_host", ""),
    }


def _query_without_page(request: HttpRequest) -> str:
    params = request.GET.copy()
    params.pop("page", None)
    return urlencode(params, doseq=True)


@login_required
def event_record_list(request: HttpRequest) -> HttpResponse:
    queryset = _event_queryset(request).prefetch_related("recording_evidences")
    paginator = Paginator(queryset, 50)
    page_obj = paginator.get_page(request.GET.get("page"))
    host_names_by_code, host_names_by_url = _inference_host_maps()

    for event in page_obj.object_list:
        event.source_host_display = _inference_host_display(
            event, host_names_by_code, host_names_by_url
        )
        event.record_snapshot_url = _snapshot_url(request, event)
        event.recording_evidence = _latest_recording(event)
        event.recording_download_url = _recording_download_url(
            request,
            event.recording_evidence,
        )
        event.recording_window = _recording_window(event.recording_evidence)
        event.recording_ready = _recording_local_mp4_ready(event.recording_evidence)
        event.recording_status_display = _recording_status_display(event.recording_evidence)
        event.recording_ui_status = _recording_ui_status(event.recording_evidence)
        event.recording_local_path = _recording_local_path(event.recording_evidence)

    areas = (
        Camera.objects.exclude(area="")
        .values_list("area", flat=True)
        .distinct()
        .order_by("area")
    )
    source_hosts = [
        {"value": host.host_code, "label": host.name or host.host_code}
        for host in InferenceHost.objects.all().order_by("name", "host_code")
    ]

    context = {
        "page_obj": page_obj,
        "record_count": paginator.count,
        "event_type_choices": Event.EVENT_TYPE_CHOICES,
        "areas": areas,
        "source_hosts": source_hosts,
        "query_without_page": _query_without_page(request),
        **_filter_context(request),
    }
    return render(request, "dashboard/event_record_list.html", context)


@login_required
def export_event_records_csv(request: HttpRequest) -> HttpResponse:
    queryset = _event_queryset(request).prefetch_related("recording_evidences")
    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="event_records_{timestamp}.csv"'
    )
    response.write("\ufeff")

    writer = csv.writer(response)
    writer.writerow(EVENT_EXPORT_HEADERS)
    for event in queryset.iterator(chunk_size=500):
        writer.writerow(_event_row(request, event))
    return response


@login_required
def export_event_records_excel(request: HttpRequest) -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        return HttpResponse(
            "缺少 openpyxl。請執行：python -m pip install openpyxl==3.1.5",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    queryset = _event_queryset(request).prefetch_related("recording_evidences")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "事件紀錄"
    worksheet.freeze_panes = "A2"
    worksheet.append(EVENT_EXPORT_HEADERS)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for event in queryset.iterator(chunk_size=500):
        worksheet.append(_event_row(request, event))

    widths = [12, 20, 22, 16, 18, 14, 22, 28, 42, 48, 18, 42, 48]
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    timestamp = timezone.localtime().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f'attachment; filename="event_records_{timestamp}.xlsx"'
    )
    return response
